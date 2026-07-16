#!/usr/bin/env python3
"""Generates AIPlatformData.xlsx seed workbook for the AI Platform dashboard."""

from pathlib import Path

try:
    from openpyxl import Workbook
except ImportError:
    raise SystemExit('Install openpyxl: pip install openpyxl')

ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / 'AIPlatformData.xlsx'

workbook = Workbook()

metrics = workbook.active
metrics.title = 'DashboardMetrics'
metrics.append(['Metric', 'Value', 'Description', 'LastUpdated'])
metrics.append(['AI Conversations', 2487, 'Total conversations', '2026-07-16'])
metrics.append(['Documents Processed', 184, 'Documents analyzed', '2026-07-16'])
metrics.append(['Automation Tasks', 96, 'Workflow success percentage', '2026-07-16'])

activity = workbook.create_sheet('ActivityLog')
activity.append(['Id', 'Timestamp', 'ActivityType', 'Title', 'Status', 'Details'])
activity.append([1, '2026-07-16 09:31', 'Document', 'Document analyzed', 'Completed', 'Contract.pdf summarized'])
activity.append([2, '2026-07-16 09:35', 'Proposal', 'Proposal generated', 'Completed', 'Software consulting proposal'])
activity.append([3, '2026-07-16 09:41', 'API', 'API documentation', 'Completed', 'REST API generated'])
activity.append([4, '2026-07-16 09:52', 'Workflow', 'Business workflow analyzed', 'Completed', 'Automation opportunities identified'])

chat = workbook.create_sheet('ChatHistory')
chat.append(['Id', 'Timestamp', 'Prompt', 'Response', 'Model', 'Tokens', 'ExecutionTime'])

workbook.save(OUTPUT)
print(f'Created {OUTPUT}')
